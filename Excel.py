'''
Created on 2017年9月3日

@author: liutao
'''
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter 
import os

class excel():



    def __init__(self, fil_address):
        self.file_address = fil_address
    
    #读取Excel表数据    
    def readExcel(self):
        """在Excel中读取数据 """
        data = []
        wb = load_workbook(self.file_address)
        sheet = wb.active
        for i in range(1, sheet.max_row+1):
            data.append(sheet.cell(row=i, column=1).value)
        wb.close()
        return data
        
    
    #写入Excel数据
    def writeExcel(self, lists):
        """把数据写入到excel中"""
        wb = Workbook()
        ws = wb.active
        #去除文件扩展名
        file_name = os.path.splitext(str(self.file_address))[0]+str('(1)')+str('.xlsx')
        print(file_name)
        print(lists)
        for i in range(0, len(lists)):
            ws.append(lists[i])
        wb.save(file_name)
        
            
        
        
        